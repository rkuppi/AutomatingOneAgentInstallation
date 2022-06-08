"""
---------------------------------------------------------
Automating the OneAgent installation on Linux
---------------------------------------------------------
Description
This script will automate the one agent installation on multiple Linux hosts which are listed in excel sheet.
All the host details are listed in the excel sheet.
This script will take the details/parameters of each host through excel, install the OneAgent at the locations given in the excell sheet and apply required configurations.
---------------------------------------------------------
INPUTS:
Excel file where all the required parameters configured with the help of Dynatarce documentation and encrypted with AES algorithm
help: https://www.dynatrace.com/support/help/setup-and-configuration/dynatrace-OneAgent/installation-and-operation/linux/installation/customize-OneAgent-installation-on-linux/
---------------------------------------------------------
OUTPUT:
Script will log all the installation steps followed.
unable_to_installList.txt: This file will be created with detailed reason for installation failure on host
successful_Installation_list.txt: Host/ServerName will be appended to this file When OneAgent services are running on the perticular host/ServerName.
---------------------------------------------------------
"""
import os
import time
import logging
# import subprocess
import paramiko
from openpyxl import load_workbook
from cryptography.fernet import Fernet, InvalidToken
import pprint
start_time = time.time()
logFilename = "OneAgentInstallation.log"
logging.basicConfig(level=logging.INFO, filename=logFilename, format='%(asctime)s %(levelname)s:%(message)s')
logging.info("-----------------------------Log file for OneAgent installation--------------------------------")
logging.info("-----Program execution was started")
logging.info("-----Processing excel sheet and loading the data into a dictionary variable")
# Note Excel file is in the same location make the necessary changes if the file is somewhere else
# all the server details are encrypted using AES algorithm use the same key to decrypt
# Decreptying the excel file using unlock.key file
try:
    logging.info("-----Decrypting the excel sheet--------------")
    with open("unlock.key", "rb") as unlock:
        key = unlock.read()
    f = Fernet(key)
    with open('enc_Configlist.xlsx', 'rb') as encrypted_file:
        encrypted = encrypted_file.read()
    decrypted = f.decrypt(encrypted)
    with open("dec_Config.xlsx", "wb") as decrypted_file:
        decrypted_file.write(decrypted)
except InvalidToken as I:
        logging.error(f"------Invalid token. Excel file cannot be decrypted") 
        exit()
except Exception as E:
        logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{E}")
        exit()
# try
try:
    logging.info("-----Loading the excel data into dictionary variable")
    workbook = load_workbook('dec_Config.xlsx')
    sheet = workbook.active
    config_data = {}
    # Note: Hardcoding the headers
    # Important Note: While creating Excel sheet please follow below naming conventions
    # headers = ["HostName", "UserName", "Password", "set-host-group", "set-host-name", "set-network-zone", "set-infra-only",	"set-proxy", "OneAgentBinaryLocation",	"OneAgentctlLocation", "DiskSizeCheck"]
    # use the new headers as required
    headers = (row for row in sheet.iter_rows(min_row=1, max_row = 1, values_only= True))
    headers = [key for tu in headers for key in tu]
    for row in sheet.iter_rows(min_row=2, values_only= True):
        row_data = [x for x in row]
        eachrow = {x:y for x,y in zip(headers, row_data)}
        config_data[row_data[0]] = eachrow
    workbook.close
    # end for
    logging.info("-------Excel data is successfully updated to a dictionary")
# end try
except NameError as N:
    logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{N}")
    os.remove("dec_Config.xlsx")
    exit()
except ValueError as V:
    logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{V}")
    os.remove("dec_Config.xlsx")
    exit()
except Exception as e:
    logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{e}")
    os.remove("dec_Config.xlsx")
    exit()
# end try except
# Removing the decrypted file
logging.info("-----Removing decrypted file")
os.remove("dec_Config.xlsx")
logging.info("------Trying to install OneAgent for each host by configuring the required parameters")
for host, parameters in config_data.items():
    print("**********************************************************************************")
    try:
        logging.info("****************************************************************")
        logging.info(f"------Trying to install OneAgent for the host {host} by configuring the required parameters")
        logging.info(f"------Trying to ping the host")
        response = os.system("ping "+ host)
        if response == 0:
            # if Response is 0, host/server is reachable
            logging.info(f"-----{host} is reachable")
            # creating a SSH session
            client = paramiko.SSHClient()
            # Auto Hiding the policy with AutoAddPolicy method
            client.set_missing_host_key_policy(paramiko.client.AutoAddPolicy())
            # Connecting with credentials
            client.connect(host,22,username=parameters["UserName"], password=parameters["Password"])
            # If no exception is raised it state that connection is successful
            logging.info(f"------Connection to the {host} was established successfully")
            logging.info(f"------Executing commands to check the disk space")
            # NOTE: Dynatrace Binary file size is 80MB in size
            # For successful installation OneAgentctl need 7.1GB data as per the documentation we are checking the free space availability of 10GB
            command = parameters["OneAgentBinaryLocation"]
            command = command.split("/")[1]
            # Ex: df /tmp
            command = f"df /{command}".rstrip()
            stdin, stdout, stderr = client.exec_command(command)
            diskspace1= stdout.readlines()
            output = ",".join(diskspace1)
            # if
            if "No such file or directory" in output:
                logging.error(f"------Command to check the memory was not found for the host {host} - {command}")
                logging.error(f"------Moving host to unable to install list")
                f = open("unable_to_installList.txt", 'a')
                f.write(f"Command {command} not found on the host - {host}\n")
                f.close()
                continue
            else:
                logging.info(f"------Command {command} to check the memory space was executed successfully ")
            # end if
            diskspace1 = int(diskspace1[1].split()[-3])
            # Converting to GB
            diskspace1 = diskspace1/(1024*1024)
            # checking Memory space for  OneAgent installation
            command = parameters["OneAgentctlLocation"]
            command = command.split("/")[1]
            command = f"df /{command}".rstrip()
            stdin, stdout, stderr = client.exec_command(command, get_pty=True)
            diskspace2 = stdout.readlines()
            output = ",".join(diskspace2)
            # if
            if "No such file or directory" in output:
                logging.error(f"------Command to check for Memory was not found for the host {host} - {command}")
                logging.error(f"------Moving host to unable to install list")
                f = open("unable_to_installList.txt", 'a')
                f.write(f"Command {command} not found on the host - {host}\n")
                f.close()
                continue
            else:
                logging.info(f"------Command {command} to check the memory space was executed successfully")
            # end if
            diskspace2 = int(diskspace2[1].split()[-3])
            # converting to GB
            diskspace2 = diskspace2/(1024*1024)
            logging.info(f"------Checking the disk space to transfer OneAgent binary file and to install it")
            # if (OneAgentBinarylocation) and (OneAgentctlLocation)
            if ((diskspace1 >= 1) and (diskspace2 >= int(parameters['DiskSizeCheck']))):
                logging.info(f"------OneAgent binary need 80MB, we have {diskspace1}GB we can transfer the file")
                logging.info(f"------To install the binary we need {parameters['DiskSizeCheck']}GB, we have {diskspace2}GB we can install the binary file transferred at the location {parameters['OneAgentctlLocation']}")
            else:
                logging.error(f"-------OneAgent binary need at least 80MB, we have {diskspace1}GB we cannot transfer the file")
                logging.error(f"-------To install the binary we need {parameters['DiskSizeCheck']}GB, we have {diskspace2} we cannot install the binary file at the location {parameters['OneAgentctlLocation']}")
                logging.error(f"-------We cannot install the OneAgent on this host {host}. We don't have required memory space")
                logging.error(f"-------Adding the host {host} to unable_to_installList.txt file")
                f = open("unable_to_installList.txt", 'a')
                f.write(f"No enough space to transfer and install the OneAgent binary for the host-{host}\n")
                f.close()
                continue
            # end if
            logging.info(f"-------Initiating  transfer and installation as we have enough space")
            # try block
            try:
                # using Secure File Transfer Protocol to transfer the OneAgent Binary to form local machine to remote server
                # Executing path exists command
                logging.info(f"-----Checking OneAgent binary path exists {parameters['OneAgentBinaryLocation']}")
                remotepath = parameters['OneAgentBinaryLocation']
                remotepath = remotepath.split("/")
                remotepath = "/".join(remotepath[0:-1])
                command = f"[ -d '{remotepath}' ] && echo 'True'".rstrip()
                stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                output = stdout.readlines()
                # if output is [] (empty)
                if len(output) == 0:
                    # Given Path does not exist
                    # Creating the directory path
                    # Ex command : mkdir -p -m777 path (Give required permissions)
                    logging.info(f"-----Path {remotepath} does not exist on the host -{host} to transfer OneAgent")
                    logging.info(f"-----Creating the {remotepath} on the host {host}")
                    # Creating a directory with all permissions, Configure this if required
                    command = f"mkdir -p -m777 {remotepath}".rstrip()
                    stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                else:
                    output = ",".join(output)
                    if "True" in output:
                        logging.info(f"------Path {remotepath} exists on the host {host} to transfer OneAgent binary")
                    # end if
                # end if
                # Please put binary file in the same directory and file name as "Dynatrace.sh"
                # creating SFTP connection between two hosts
                sftp = client.open_sftp()
                remotepath = parameters['OneAgentBinaryLocation']
                response = sftp.put("Dynatrace.sh", remotepath)
                sftp.close()
                # We receive response for the above transfer
                # if block
                if response:
                    logging.info(f"------OneAgent Binary file was transferred to {parameters['OneAgentBinaryLocation']}")
                    logging.info(f"------Changing the file permissions for {parameters['OneAgentBinaryLocation']}")
                    command = parameters['OneAgentBinaryLocation']
                    # changing the file permission
                    # "chmod 755" + " " + command
                    command = f"chmod 755 {command}".rstrip()
                    stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                    output = stdout.readlines()
                    if len(output) == 0:
                        logging.info(f"------File permissions are changed successfully") 
                    else:
                        output = ",".join(output)
                        logging.error(f"------Output to the command -{command}-{output}")
                        logging.error(f"------Command not found for the host {host} - {command}")
                        logging.error(f"------Moving host to unable to install list")
                        f = open("unable_to_installList.txt", 'a')
                        f.write(f"Command {command} not found on the host-{host}\n")
                        f.close()
                        continue
                    # end else
                    # command = "sudo -S "+ "/bin/sh Dynatrace-OneAgent-AIX-1.0.0.sh INSTALL_PATH=/data/dynatrace/agent"
                    # Follow the documentation https://www.dynatrace.com/support/help/setup-and-configuration/dynatrace-OneAgent/installation-and-operation/linux/installation/customize-OneAgent-installation-on-linux/
                    # Specify the custom installation path in excel sheet.
                    # checking the file path
                    logging.info(f"------Checking path to install OneAgent")

                    # Installing OneAgent
                    command = parameters["OneAgentBinaryLocation"]
                    # "sudo -S"+ " " + command
                    logging.info(f"-----Installing the binary file at the location -{parameters['OneAgentctlLocation']}")
                    command = f"sudo -S {command}".rstrip()
                    stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                    stdin.write(parameters["Password"]+ "\n")
                    stdin.flush()
                    output = stdout.readlines()
                    pprint.pprint(output)
                    logging.info(f"-----Output of installation\n{output}")
                    output = ",".join(output)
                    # NOTE: Capturing Error statements here
                    if ("Error" in output) or ("error" in output) or ("For details, see" in output):
                        logging.error(f"------Error occurred on the host {host} while installing OneAgent with the command- {command}")
                        command = "systemctl | grep Dynatrace"
                        stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                        output = stdout.readlines()
                        output.append("Dummy element")
                        output = ",".join(output)
                        if "running" in output:
                            logging.error(f"------OneAgent is running of the host but there are error logs during installation on host -{host}")
                        # end if
                        else:
                            logging.info(f"------One agent is not running on the host - {host} Error occurred during installation")
                            f = open("unable_to_installList.txt", 'a')
                            f.write(f"Error occurred while installing on the host-{host}\n")
                            f.close()
                            continue
                        # end else
                    # end if
                    else:
                        logging.info(f"------OneAgent was installed on host {host}")
                    # end if
                    logging.info(f"------Configuring OneAgent parameters")
                    # Configuring the parameters
                    selection = ['set-host-group', 'set-host-name', 'set-network-zone', 'set-infra-only', 'set-proxy']
                    filtered = dict(filter(lambda i:i[0] in selection, parameters.items()))
                    # Following this format:--set-host-group=%hostgroup% --set-infra-only=true --set-network-zone=%networkzone% --set-host-name=%Customhostname% --set-proxy=%proxy% --restart-service
                    config_command_string = ""
                    for key, value in filtered.items():
                        if value == None:
                            logging.info(f"-----Value for the parameter {key}={value} cannot configure {key} for the host {host}")
                            continue
                        # end if
                        elif (value == "-") or (value == "None"):
                            logging.info(f"-----Value for the parameter {key}={value} cannot configure {key} for the host {host}")
                            continue
                        # end elif
                        else:
                            logging.info(f"-----Value for the parameter {key}={value} configuring {key} for the host {host}")
                            config_command_string = config_command_string + f"--{key}={value}" + " "
                        # end else
                    config_command_string= "sudo -S" + " " + f"{parameters['OneAgentctlLocation']}" + " " + config_command_string + "--restart-service"
                    # end of for
                    # executing the config_command_string
                    logging.info(f"-----Executing {config_command_string} command on the host {host}")
                    stdin, stdout, stderr = client.exec_command(config_command_string.rstrip(), get_pty=True)
                    stdin.write(parameters["Password"]+ "\n")
                    stdin.flush()
                    logging.info(f"------Output of the executed command\n{stdout.readlines()}")
                    logging.info(f"-----Checking the OneAgent services running on the host {host} to confirm installation")
                    command = "systemctl | grep Dynatrace"
                    stdin, stdout, stderr = client.exec_command(command.rstrip(), get_pty=True)
                    output = stdout.readlines()
                    # Adding dummy element to avoid exception when the output is empty list
                    output.append("Dummy element")
                    output = ",".join(output)
                    if "running" in output:
                        logging.info(f"-------OneAgent was installed successfully and OneAgent service is running on the host {host}")
                        pprint.pprint(f"OneAgent installation was successful on the {host}")
                        logging.info(f"-----Adding host to successful_Installation_list.txt")
                        f = open("successful_Installation_list.txt", "a+")
                        f.write(f"{host}\n")
                        f.close
                    else:
                        logging.error(f"-------Cross check the installation and configuration parameters applied for the host-{host}")
                        logging.error(f"-------OneAgent service is not running on the host {host}")
                        logging.error(f"------Adding host {host} to unable to install list")
                        f = open("unable_to_installList.txt", 'a')
                        f.write(f"OneAgent services are not running on the host {host}\n")
                        f.close
                    # end of else (checking OneAgent.service status)
                # end of if statement (check for binary file is transferred to the remote location)
                else:
                    logging.error(f"------OneAgent file is not transferred to {parameters['OneAgentBinaryLocation']}")
                    logging.error(f"------Cannot intall OneAge on this host {host}, check the admin or group permissions for the folder/file {parameters['OneAgentBinaryLocation']}")
                    logging.error(f"------Adding the host {host} to unable_to_installList")
                    f = open("unable_to_installList.txt", 'a')
                    f.write(f"File is not transferred cannot install agent on the host -{host}\n")
                    f.close
            # end of try statement (To cover Binary file transfer and installation exceptions)
            except paramiko.SFTPError as S:
                logging.error(f"-----SFTP Error occurred  {host} Error-{A}")
                logging.error(f"-----Adding to unable_to_installList")
                f = open("unable_to_installList.txt", 'a')
                f.write(f"SFTP Error -{S} on the host-{host}\n")
                f.close
                if client:
                    logging.info(f"------Closing the ssh session for the host {host}")
                # client.close()
            except Exception as E:
                logging.error(f"Exception -{E}")
                logging.error(f"-----Exception occurred on the host {host} Error-{E}")
                logging.error(f"-----Adding to unable_to_installList")
                f = open("unable_to_installList.txt", 'a')
                f.write(f"Exception {E} On the host-{host}\n")
                f.close
                if client:
                    logging.info(f"------Closing the ssh session for the host {host}")
                # client.close()
            # closing the SSH session
            # end of except statement (To cover Binary file transfer and installation exceptions)
            client.close()
        else:
            logging.error(f"------Cannot reach the host {host}, we cannot proceed for the installation of OneAgent")
            logging.error(f"------Adding the host {host} to unable_to_installList")
            f = open("unable_to_installList.txt", 'a')
            f.write(f"Unable to reach host-{host}\n")
            f.close()
        # end of ping test case
    # end of outer try block (ping the network) 
    except paramiko.AuthenticationException as A:
        logging.error(f"-----Authentication Error for the host {host} Error-{A}")
        logging.error(f"-----Adding to unable_to_installList")
        f = open("unable_to_installList.txt", 'a')
        f.write(f"Authentication Exception {A} for the host-{host}\n")
        f.close
        if client:
            logging.info(f"------Closing the ssh session for the host {host}")
    except paramiko.SSHException as S:
        logging.error(f"-----SSH Error for the host {host} Error-{S}")
        logging.error(f"-----Adding to unable_to_installList")
        f = open("unable_to_installList.txt", 'a')
        f.write(f"SSH Exception {S} for the host {host} \n")
        f.close
        if client:
            logging.info(f"------Closing the ssh session for the host {host}")
    except Exception as E:
        logging.error(f"-----Exception occurred for the host {host} -{E}")
        logging.error(f"-----Adding to unable_to_installList")
        f = open("unable_to_installList.txt", 'a')
        f.write(f"Exception occurred- Exception {E} for the host - {host}\n")
        f.close
        if client:
            logging.info(f"------Closing the ssh session for the host {host}")
    print("*********************************************************************")
    logging.info("*************************************************************************************************")
    # ending for loop
# end for
print("Execution  finished in --- %s min ---" % ((time.time() - start_time)/60))