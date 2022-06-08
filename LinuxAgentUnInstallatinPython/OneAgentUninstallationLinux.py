"""
--------------------------------------------------------------------------------------------------
Script will uninstall the OneAgents on the server and clean up respective directory paths in Linux
--------------------------------------------------------------------------------------------------

Input: An excel file with servers details mentioned, use the template to fill the details
Output: Clean and uninstall OneAgents
"""
import paramiko
from openpyxl import load_workbook
import logging
import os
from cryptography.fernet import Fernet
logFilename = "OneAgentUninstallation.log"
logging.basicConfig(level=logging.INFO, filename=logFilename, format='%(asctime)s %(levelname)s:%(message)s')
try:
    logging.info("------Decrypting the excel sheet")
    #  adjust the encrypted file accordingly
    with open("unlock.key", "rb") as unlock:
        key = unlock.read()
    f = Fernet(key)
    with open('enc_UnstallationList.xlsx', 'rb') as encrypted_file:
        encrypted = encrypted_file.read()
    decrypted = f.decrypt(encrypted)
    with open("dec_UninstallationList.xlsx", "wb") as decrypted_file:
        decrypted_file.write(decrypted)
except Exception as E:
        logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{E}")
        exit()
# try
try:
    logging.info("-----Loading the excel data into dictionary variable")
    workbook = load_workbook('dec_UninstallationList.xlsx')
    sheet = workbook.active
    config_data = {}
    # Note: Hardcoding the headers
    # Important Note: While creating Excel sheet please follow below naming conventions
    # headers = [HostName, UserName, Password, OneAgentBinaryLocation, OneAgentctlLocation, RemoveLogs, Remove DynatraceFolder]
    headers = (row for row in sheet.iter_rows(min_row=1, max_row = 1, values_only= True))
    headers = [key for tu in headers for key in tu]
    for row in sheet.iter_rows(min_row=2, values_only= True):
        row_data = [x for x in row]
        eachrow = {x:y for x,y in zip(headers, row_data)}
        config_data[row_data[0]] = eachrow
    workbook.close
    logging.info("-------Excel data is successfully updated to a dictionary")
except NameError as N:
    logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{N}")
    os.remove("dec_UninstallationList.xlsx")
    exit()
except ValueError as V:
    logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{V}")
    os.remove("dec_UninstallationList.xlsx")
    exit()
except Exception as e:
    logging.error(f"-----Occurred while processing the the excel sheet-ERROR -{e}")
    os.remove("dec_UninstallationList.xlsx")
    exit()
os.remove("dec_UninstallationList.xlsx")
#  for each host
for host, parameters in config_data.items():
    print("**********************************************************************************")
    try:
        logging.info("****************************************************************")
        logging.info(f"------Trying to uninstall OneAgent for the host {host} by configuring the required parameters")
        logging.info(f"------Trying to ping the host")
        response = os.system("ping "+ host)
        if response == 0:
            # if Response is 0, host/server is reachable
            logging.info(f"-----{host} is reachable")
            # creating a SSH session
            client = paramiko.SSHClient()
            # Auto Hiding the policy with AutoAddPolicy method
            client.set_missing_host_key_policy(paramiko.client.AutoAddPolicy())
            # Connecting with credentials
            client.connect(host,22,username=parameters["UserName"], password=parameters["Password"])
            # If no exception is raised it state that connection is successful
            logging.info(f"------Connection to the {host} was established successfully")
            # Executing Un-installation commands
            # running Dynatrace uninstall.sh file
            command = 'sudo find / -type f -name "uninstall.sh"'
            stdin, stdout, stderr = client.exec_command(command, get_pty=True)
            stdin.write(parameters["Password"]+ "\n")
            stdin.flush()
            output = stdout.readlines()
            if len(output) == 0:
                logging.info(f"-----There are no uninstall.sh found on entire system({host})")
            else:
                # checking for /OneAgent/agent/uninstall.sh
                for path in output:
                    if "/OneAgent/agent/uninstall.sh" in path:
                        logging.info(f"-----Dynatrace uninstall.sh exists on the host -{host}")
                        logging.info(f"------Executing Dynatrace uninstall.sh file on the host -{host}")
                        uninstallation_command = f"sudo -S {path}".rstrip()
                        stdin, stdout, stderr = client.exec_command(uninstallation_command, get_pty=True)
                        stdin.write(parameters["Password"]+ "\n")
                        stdin.flush()
                        output = stdout.readlines()
                        logging.info(f"------Output of uninstall command -{output}")
                    # end if
                # end for
            # end else
            # removing config files of Dynatrace
            logging.info(f"-----Removing config files related to Dynatrace")
            command = "sudo find / -type d -name 'config'"
            stdin, stdout, stderr = client.exec_command(command, get_pty=True)
            stdin.write(parameters["Password"]+ "\n")
            stdin.flush()
            output = stdout.readlines()
            if len(output) == 0:
                logging.info(f"-----There are no config folders on entire system({host})")
            else:
                # checking for /OneAgent/agent/uninstall.sh
                for path in output:
                    if "/OneAgent/agent/config" in path:
                        logging.info(f"-----Dynatrace config files exists on the host -{host}")
                        uninstallation_command = f"sudo -S rm -r {path}".rstrip()
                        stdin, stdout, stderr = client.exec_command(uninstallation_command, get_pty=True)
                        stdin.write(parameters["Password"]+ "\n")
                        stdin.flush()
                        output = stdout.readlines()
                        logging.info(f"------Output of removing confing files command -{output}")
                    # end if
                # end for
            #  end if
            logging.info(f"------Removing Dynatrace logs on the host -{host} with user input in from the excel")
            # Removing logs folder based on excel sheet data
            remove = parameters["RemoveLogs"]
            if (remove == "Yes") or (remove == "yes"):
                command = "sudo find / -type d -name 'log'"
                stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                stdin.write(parameters["Password"]+ "\n")
                stdin.flush()
                output = stdout.readlines()
                if len(output) == 0:
                    logging.info(f"-----There are no log folders on entire system({host})")
                else:
                    for path in output:
                        if "/OneAgent/log" in path:
                            logging.info(f"-----Dynatrace log files exists on the host -{host}")
                            logging,logging.info(f"------Executing command to remove dynatarce logs file on the host -{host}")
                            uninstallation_command = f"sudo -S rm -r {path}".rstrip()
                            stdin, stdout, stderr = client.exec_command(uninstallation_command, get_pty=True)
                            stdin.write(parameters["Password"]+ "\n")
                            stdin.flush()
                            output = stdout.readlines()
                            logging.info(f"------Output of removing log files command -{output}")
                        # end if
                    # end for
                # end else
            # end if
            else:
                logging.info(f"-----Not removing log the host -{host}")
            logging.info(f"------Removing the OneAgent binary file {parameters['OneAgentBinaryLocation']} on the host {host}")
            # removing OneAgent binary which was transferred through script
            # Note: need to give full path
            location = parameters['OneAgentBinaryLocation']
            uninstallation_command = f"sudo -S rm -r {location}".rstrip()
            stdin, stdout, stderr = client.exec_command(uninstallation_command, get_pty=True)
            stdin.write(parameters["Password"]+ "\n")
            stdin.flush()
            output = stdout.readlines()
            logging.info(f"-------Output of removing binaryfile command-{output}")
            logging.info(f"-----Removing Dynatrace folder based on user input in excel sheet for the host {host}")
            # Removing Dynatrace folder based on user input in excel sheet
            remove = parameters["Remove DynatraceFolder"]
            if (remove == "Yes") or (remove == "yes"):
                location = parameters["OneAgentctlLocation"]
                location = location.split("/")
                index = location.index("Dynatrace")
                command = "/".join(location[1:index+1])
                command = f"sudo -S rm -r {command}".rstrip()
                stdin, stdout, stderr = client.exec_command(command, get_pty=True)
                stdin.write(parameters["Password"]+ "\n")
                stdin.flush()
                output = stdout.readlines()
                logging.info(f"------Output of removing Dynatrace folder command -{output}")
            # end if
            else:
                logging.info(f"-----Not removing Dynatrace folder")
            client.close()
        # end if
        else:
            logging.error(f"------Host is not reachable-{host}")
    except paramiko.AuthenticationException as A:
        logging.error(f"-----Authentication Error for the host {host} Error-{A}")
        if client:
            logging.info(f"------Closing the ssh session for the host {host}")
    except paramiko.SSHException as S:
        logging.error(f"-----Adding to unable_to_installList")
        if client:
            logging.info(f"------Closing the ssh session for the host {host}")
    except Exception as E:
        logging.error(f"-----Exception occurred for the host {host} -{E}")
        if client:
            logging.info(f"------Closing the ssh session for the host {host}")
    # end try
# end for